import json
import os
import zipfile
import openpyxl
import gdown

url = "https://docs.google.com/spreadsheets/d/1Yvgjex502cB_dVvvZm0a88aGL4WNFOm-5XvEbZLkWqI/export"
workbookPath = "dj_datasheet.xlsx"

def processExcel():
  try:
    # Download the datasheet
    gdown.download(url, workbookPath, fuzzy=True, format="xlsx")

    workbook = openpyxl.load_workbook(workbookPath)

    # Collect the full names of the converters
    print("------ GETTING CONVERTER INFORMATION ------")
    converters = {}
    informationSheet = workbook["Information"]
    for row in range(2, informationSheet.max_row):
      abbrName = informationSheet.cell(row=row, column=1).value
      fullName = informationSheet.cell(row=row, column=2).value

      # Skip any null values
      if(abbrName is None or fullName is None): continue
      converters[abbrName] = fullName
      print("Found converter: " + abbrName + " - " + fullName)

      #for col in informationSheet.iter_cols(1, 2):
      #  print(col[row].value)


    # Now process the worksheets!
    sheetName = "GO BGM Vanilla"
    #sheetName = "DJ BGM Vanilla"
    sheet = workbook[sheetName]
    print("------ PROCESSING SHEET: " + sheetName + " ------")
    

    # We allow an offset, because some sheets can have 2 tables of data
    for offset in [0, 6]:
      for row in range(2, sheet.max_row):
          game = sheet.cell(row=row, column=2+offset).value
          title = sheet.cell(row=row, column=3+offset).value
          progress = sheet.cell(row=row, column=4+offset).value
          author = sheet.cell(row=row, column=5+offset).value or ""
          sample = sheet.cell(row=row, column=6+offset).value or ""

          # If a track is WIP, we skip it
          if(game is None or title is None or not isinstance(title, str) or progress != "Done"): continue

          # If a track is a boolean, then it's probably because the name of the song is "True" or "False"
          if(isinstance(title, bool)): title = "True" if title else "False"

          print("Found song: " + game + " - " + title + " | " + author + " | " + sample)
    
    # If everything when ok, then return true
    return True
  
  # On exceptions, always print them
  except Exception as e:
    print(e)
    return False
  
  # Always try to delete the datasheet, so no evindence remains
  finally:
    if os.path.exists(workbookPath): os.remove(workbookPath)


def main():
  result = processExcel()
  if result: print("Process completed succesfully!")
  else: print("An error occured")


if __name__ == '__main__':
  main()